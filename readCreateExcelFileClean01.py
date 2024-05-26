from openpyxl import load_workbook

from datetime import date
today = date.today()

# load the excel
wb = load_workbook(filename = "Raw_data_Clients_Savings_2024_05_20.xlsx")

# grab the active Sheet in the excel
activeWorkSheet = wb.active

# Create an empty list to store the data from column A
column_a_data = []

# Iterate through each row in column A and append the value to the list
for row in activeWorkSheet['A2:A100']:
    for cell in row:
        if cell.value is not None:
            column_a_data.append(cell.value)
        else:
            # Handle empty cells, as per your requirements
            column_a_data.append(None)  # Placeholder value or None, based on your preference

# Create an empty list to store the data from column B
column_b_data = []

for row in activeWorkSheet['B2:B100']:
    for cell in row:
        if cell.value is not None:
            column_b_data.append(cell.value)
        else:
            column_b_data.append(None)  # Placeholder value or None, based on your preference

# combine column a list and column b list as 1 list
zipValueListAListB = zip(column_a_data,column_b_data)
listZipValueListAListB = list(zipValueListAListB)

# to create a group a list of Portfolios which contain less or equal 5k transactions
column_c_data = []
total450k = 0
groupBlock = 1
for valueA, valueB in listZipValueListAListB:
    total450k = valueB + total450k
    if total450k <= 450000:
        groupBlock = groupBlock
        column_c_data.append(groupBlock)
    else:
        groupBlock = groupBlock + 1
        column_c_data.append(groupBlock)
        total450k = valueB

# to save the above 3 lists; column_a_data, column_b_data, and column_c_data into new excel
from openpyxl import Workbook

# Create a new workbook
newWorkbookExcel = Workbook()

# Select the active worksheet in the new workbook
newWorkbookWorksheetExcel = newWorkbookExcel.active

# transfer list to column
for index, itemInCellColumnA in enumerate(column_a_data):
    newIndex = index + 1
    newWorkbookWorksheetExcel.cell(row=newIndex + 1, column=1, value=itemInCellColumnA)

for index, itemInCellColumnB in enumerate(column_b_data):
    newIndex = index + 1
    newWorkbookWorksheetExcel.cell(row=newIndex + 1, column=2, value=itemInCellColumnB)

for index, itemInCellColumnC in enumerate(column_c_data):
    newIndex = index + 1
    newWorkbookWorksheetExcel.cell(row=newIndex + 1, column=3, value=itemInCellColumnC)

# add headers to excel
newWorkbookWorksheetExcel.cell(row=1, column=1).value="Client"
newWorkbookWorksheetExcel.cell(row=1, column=2).value="Saving"
newWorkbookWorksheetExcel.cell(row=1, column=3).value="Block/Group"

# Save the workbook --> create the new excel
newWorkbookExcel.save('new_Clients_Savings_Group_' + str(today) + '.xlsx')

# Close the old Excel file
wb.close()

# Close the new Excel file
newWorkbookExcel.close()